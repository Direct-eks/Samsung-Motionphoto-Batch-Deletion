import binascii
import os
from openpyxl import Workbook


def folder_process(path, number, path_to):

    if path == "skip":
        return

    files = os.listdir(path)
    number_of_files = files.__len__()

    print("There are ", end='')
    print(number_of_files, end='')
    print(" files in the directory")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Processed"
    # worksheet initialization
    for row in range(1, number_of_files):
        for col in range(1,4):
            worksheet.cell(row, col, "null")

    row = 1

    Mo_binary = binascii.a2b_qp('MotionPhoto_Data')

    for filename in files:
        # check if filename is a directory
        if os.path.isdir(filename):
            continue

        # generate filename for processed file
        target_name = path + '/' + filename
        write_name = path_to + '/' + filename

        # open file for process
        with open(target_name, 'rb') as file:
            # total byte length of the file
            length = os.path.getsize(target_name)

            worksheet.cell(row, 1, filename)
            worksheet.cell(row, 2, float(length / 1000000))

            if length < 1000 * 1000 + 500 * 1000:
                print(filename, end='')
                print("size < 1.5MB")
                continue

            file.seek(-200, 2)
            content = file.read(200)
            already_cut = False

            index = content.rfind(b'\xFF\xD9')
            if index != -1:
                already_cut = True

            if not already_cut:
                file.seek(0, 0)
                jpg_content = file.read(1500 * 1000)
                counter = 1500 * 1000

                leftover_size = length - counter

                while leftover_size >= 200 * 1000:

                    content = file.read(200 * 1000)
                    counter += 200 * 1000

                    index = content.rfind(Mo_binary)
                    if index != -1:
                        file.seek(-200 * 1000, 1)
                        jpg_content += file.read(index + 1)

                        worksheet.cell(row, 3, float((counter - 200 * 1000 + index + 1) / 1000000) )

                        with open(write_name, 'wb') as new_file:
                            new_file.write(jpg_content)
                            new_file.close()
                        break
                    else:
                        file.seek(-200 * 1000, 1)
                        jpg_content += file.read(200 * 1000 - 20)
                        counter -= 20
                        leftover_size = length - counter
                # end while

            file.close()
        # with file open() end

        print(number, end='')
        print(' Total process: ', end='')
        print(float(row / number_of_files * 100), end='')
        print(" percent")
        row += 1

    workbook_name = 'I:/S8/process_data' + str(number) + '.xlsx'
    workbook.save(workbook_name)
    print(number, end='')
    print(' done')

# end folder_process()
